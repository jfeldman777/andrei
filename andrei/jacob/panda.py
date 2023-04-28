import pandas as pd

data = [[1,2,3],[4,5,6],[7,8,9]]


df = pd.DataFrame(data)


from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# Write column headers
for col_num, column_title in enumerate(df.columns, 1):
    ws.cell(row=1, column=col_num, value=column_title)

# Write data rows
for row_num, row_data in df.iterrows():
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num+1, column=col_num, value=cell_value)



import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from tkinter import ttk

class ExcelWidget(tk.Frame):
    def __init__(self, master, label):
        super().__init__(master)
        self.label = label
        self.filename = tk.StringVar()
        self.sheetname = tk.StringVar()
        self.headers = []
        self.data = []
        self.table = ttk.Treeview(self)
        self.table['columns'] = []
        self.table['show'] = 'headings'
        self.table.grid(row=10, column=10, columnspan=3)
        tk.Label(self, text=label).grid(row=0, column=0)
        tk.Entry(self, textvariable=self.filename, width=40).grid(row=0, column=1)
        tk.Button(self, text='Browse', command=self.browse_file).grid(row=0, column=2)
        tk.Label(self, text='Sheet Name:').grid(row=2, column=0)
        tk.Entry(self, textvariable=self.sheetname).grid(row=2, column=1)
        tk.Button(self, text='Load', command=self.load_data).grid(row=2, column=2)

    def browse_file(self):
        self.filename.set(filedialog.askopenfilename())

    def load_data(self):
        wb = load_workbook(filename=self.filename.get())
        sheet = wb[self.sheetname.get()]
        self.headers = [cell.value for cell in sheet[1]]
        self.table['columns'] = self.headers
        self.table.delete(*self.table.get_children())
        for row in sheet.iter_rows(min_row=2):
            data_row = [cell.value for cell in row]
            self.table.insert('', tk.END, values=data_row)
        for col in sheet.columns:
            col_letter = get_column_letter(col[0].column)
            width = max([len(str(cell.value)) for cell in col])
            self.table.column(col_letter, width=width * 7)
            self.table.heading(col_letter, text=col[0].value)

class Application(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        # self.left_top_widget = ExcelWidget(self, 'Left top')
        # self.right_top_widget = ExcelWidget(self, 'Right top')
        # self.left_bottom_widget = ExcelWidget(self, 'Left bottom')
        self.right_bottom_widget = ExcelWidget(self, 'Right bottom')
        # self.left_top_widget.grid(row=0, column=0)
        # self.right_top_widget.grid(row=0, column=1)
        # self.left_bottom_widget.grid(row=1, column=0)
        self.right_bottom_widget.grid(row=1, column=1)

if __name__ == '__main__':
    root = tk.Tk()
    root.title('Excel Widgets')
    app = Application(root)
    app.grid()
    root.mainloop()

# Save workbook
#wb.save('output.xlsx')
